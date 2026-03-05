import asyncio
import json
import logging
import os
import sys
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
from urllib.parse import urlencode
import httpx
import feedparser
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, stream=sys.stdout)
logger = logging.getLogger(__name__)

class PodcastIntelligencePro:
    def __init__(self):
        self.itunes_base_url = "https://itunes.apple.com"
        self.client = None
        
    async def initialize(self):
        self.client = httpx.AsyncClient(timeout=30.0)
        
    async def cleanup(self):
        if self.client:
            await self.client.aclose()
    
    async def search_itunes(self, search_query: str, country: str = "US",
                           language: Optional[str] = None, genre_id: Optional[int] = None,
                           max_results: int = 50) -> List[Dict[str, Any]]:
        try:
            params = {
                "term": search_query,
                "country": country,
                "media": "podcast",
                "limit": min(max_results, 200),
                "entity": "podcast"
            }
            
            if language:
                params["lang"] = language
            
            url = f"{self.itunes_base_url}/search?{urlencode(params)}"
            response = await self.client.get(url)
            response.raise_for_status()
            data = response.json()
            
            podcasts = []
            for result in data.get("results", []):
                if result.get("kind") == "podcast":
                    podcast = {
                        "itunes_id": result.get("collectionId"),
                        "title": result.get("collectionName"),
                        "artist": result.get("artistName"),
                        "description": result.get("description", ""),
                        "artwork_url": result.get("artworkUrl600"),
                        "feed_url": result.get("feedUrl"),
                        "country": country,
                        "primary_genre": result.get("primaryGenreName"),
                        "track_count": result.get("trackCount", 0),
                        "release_date": result.get("releaseDate"),
                        "itunes_url": result.get("collectionViewUrl"),
                        "host_name": result.get("artistName"),
                    }
                    podcasts.append(podcast)
            
            logger.info(f"Found {len(podcasts)} podcasts on iTunes")
            return podcasts
        except Exception as e:
            logger.error(f"Error searching iTunes: {str(e)}")
            return []
    
    async def parse_rss_feed(self, feed_url: str, max_episodes: Optional[int] = None) -> Dict[str, Any]:
        try:
            response = await self.client.get(feed_url, follow_redirects=True, timeout=15)
            response.raise_for_status()
            feed = feedparser.parse(response.text)
            
            if not feed.get("feed"):
                return {}
            
            feed_data = feed["feed"]
            episodes = []
            for entry in feed.get("entries", [])[:max_episodes or None]:
                episode = {
                    "title": entry.get("title", ""),
                    "description": entry.get("description", ""),
                    "link": entry.get("link", ""),
                    "published": entry.get("published", ""),
                }
                episodes.append(episode)
            
            return {"episodes": episodes, "total_episodes": len(episodes)}
        except Exception as e:
            return {"episodes": [], "total_episodes": 0}
    
    async def scrape_podcast_website(self, podcast_url: str) -> Dict[str, Any]:
        try:
            if not podcast_url or not podcast_url.startswith("http"):
                return {}
            
            response = await self.client.get(podcast_url, follow_redirects=True, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, "html.parser")
            text = soup.get_text()
            
            contact_info = {}
            
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            emails = re.findall(email_pattern, text)
            
            if emails:
                contact_info["contact_email"] = emails[0]
            
            mailto_links = soup.find_all("a", href=re.compile(r"^mailto:"))
            if mailto_links:
                email = mailto_links[0].get("href").replace("mailto:", "").split("?")[0]
                if email and "@" in email:
                    contact_info["host_email"] = email
            
            name_pattern = r'(?:Hosted by|Host|Creator|Created by)\s+([A-Z][a-z]+ (?:[A-Z][a-z]+)?)'
            names = re.findall(name_pattern, text)
            if names:
                contact_info["contact_name"] = names[0]
            
            return contact_info
        except Exception as e:
            return {}
    
    async def enrich_podcast_data(self, podcast: Dict[str, Any], include_episodes: bool = True,
                                  max_episodes: Optional[int] = None) -> Dict[str, Any]:
        enriched = podcast.copy()
        
        if podcast.get("feed_url") and include_episodes:
            rss_result = await self.parse_rss_feed(podcast["feed_url"], max_episodes)
            enriched["total_episodes"] = rss_result.get("total_episodes", 0)
        
        if podcast.get("itunes_url"):
            website_data = await self.scrape_podcast_website(podcast.get("itunes_url"))
            enriched.update(website_data)
        
        enriched["extracted_at"] = datetime.utcnow().isoformat()
        return enriched
    
    async def main(self, input_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        mode = input_data.get("mode", "search")
        podcasts = []
        
        if mode == "search":
            podcasts = await self.search_itunes(
                search_query=input_data.get("searchQuery", ""),
                country=input_data.get("country", "US"),
                language=input_data.get("language"),
                genre_id=input_data.get("genreId"),
                max_results=input_data.get("maxResults", 50)
            )
        
        enriched_podcasts = []
        for podcast in podcasts:
            enriched = await self.enrich_podcast_data(
                podcast,
                include_episodes=input_data.get("includeEpisodes", True),
                max_episodes=input_data.get("maxEpisodesPerPodcast")
            )
            enriched_podcasts.append(enriched)
        
        logger.info(f"Total podcasts enriched: {len(enriched_podcasts)}")
        return enriched_podcasts


def create_excel(results: List[Dict[str, Any]], filename: str):
    """Create Excel file with podcast data"""
    logger.info(f"Creating Excel file: {filename}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Podcasts"
    
    headers = [
        "iTunes ID", "Title", "Host Name", "Host Email", "Contact Name",
        "Contact Email", "Artist", "Description", "Genre", "Episode Count",
        "Track Count", "Release Date", "Country", "Feed URL", "iTunes URL",
        "Total Episodes", "Extracted At"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    logger.info(f"Adding {len(results)} rows of data")
    
    for row_num, result in enumerate(results, 2):
        ws.cell(row=row_num, column=1).value = result.get("itunes_id", "")
        ws.cell(row=row_num, column=2).value = result.get("title", "")
        ws.cell(row=row_num, column=3).value = result.get("host_name", "")
        ws.cell(row=row_num, column=4).value = result.get("host_email", "")
        ws.cell(row=row_num, column=5).value = result.get("contact_name", "")
        ws.cell(row=row_num, column=6).value = result.get("contact_email", "")
        ws.cell(row=row_num, column=7).value = result.get("artist", "")
        ws.cell(row=row_num, column=8).value = result.get("description", "")
        ws.cell(row=row_num, column=9).value = result.get("primary_genre", "")
        ws.cell(row=row_num, column=10).value = result.get("total_episodes", 0)
        ws.cell(row=row_num, column=11).value = result.get("track_count", 0)
        ws.cell(row=row_num, column=12).value = result.get("release_date", "")
        ws.cell(row=row_num, column=13).value = result.get("country", "")
        ws.cell(row=row_num, column=14).value = result.get("feed_url", "")
        ws.cell(row=row_num, column=15).value = result.get("itunes_url", "")
        ws.cell(row=row_num, column=16).value = result.get("total_episodes", 0)
        ws.cell(row=row_num, column=17).value = result.get("extracted_at", "")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
        ws.column_dimensions[col].width = 20
    
    try:
        wb.save(filename)
        logger.info(f"Excel file saved successfully: {filename}")
        
        if os.path.exists(filename):
            file_size = os.path.getsize(filename)
            logger.info(f"File exists. Size: {file_size} bytes")
        else:
            logger.error(f"File was not created: {filename}")
    except Exception as e:
        logger.error(f"Error saving Excel file: {str(e)}", exc_info=True)


async def run():
    input_data = {}
    input_file = os.getenv("APIFY_INPUT_FILE")
    
    if input_file and os.path.exists(input_file):
        with open(input_file, 'r') as f:
            input_data = json.load(f)
    
    if not input_data:
        input_data = {"mode": "search", "searchQuery": "technology", "country": "US", "maxResults": 5, "includeEpisodes": True}
    
    logger.info(f"Starting with input: {json.dumps(input_data)}")
    
    intelligence = PodcastIntelligencePro()
    await intelligence.initialize()
    
    try:
        results = await intelligence.main(input_data)
        logger.info(f"Successfully processed {len(results)} podcasts")
        
        if results:
            excel_filename = "podcast_results.xlsx"
            create_excel(results, excel_filename)
            logger.info(f"Excel export complete")
        else:
            logger.warning("No results to export")
            
    except Exception as e:
        logger.error(f"Fatal error: {str(e)}", exc_info=True)
    finally:
        await intelligence.cleanup()


if __name__ == "__main__":
    asyncio.run(run())
